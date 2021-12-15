# -*- coding: utf-8 -*-
"""
Created on Wed Dec 15 22:55:12 2021

@author: 皮皮卡卡
"""
from bs4 import BeautifulSoup as BS
import requests
import datetime
import re
import openpyxl as pyxl
import time as t

'''dictionary of station and number'''

wb = pyxl.load_workbook('stationtonumber.xlsx')
sheet = wb['worksheet1']
sheet = wb.active
station_number = {}

for i in range(1,sheet.max_row,2):
    #station_number.setdefault(sheet.cell(row=i, column=1).value,
    #                          sheet.cell(row=i+1, column=1).value)
    station_number[sheet.cell(row=i, column=1).value] = sheet.cell(row=i+1, column=1).value   

'''get input start-station and end-station and them numbers'''

st = input('Enter the departure station:')
ed = input('Enter the arrival station:')
start = t.time()
def getstations_input(station_name):
    a = list(station_name)
    if a[0] == '台':
        a[0] = '臺'
    a = ''.join([i for i in a]) #必須是臺這個才能查
    a = str(station_number[a]) + '-'  + a

    return a

'''comfirm the data that will be inputted'''

dt = datetime.datetime.now()
day = dt.strftime('%Y/%m/%d')
Min = int(dt.strftime('%M'))
hour = int(dt.strftime('%H'))
if Min >= 30:
    hour +=1
    Min = '00'
else:
    Min = '00'

time = str(hour)+':'+Min    
end = str(hour+1) + ':' +Min
if hour == 23:
    end = '23:59'
    
data = {
        '_csrf': '1275647f-8532-4d14-9fb0-bcb84bf61f4c',
        'startStation': "",
        'endStation': "",
        'transfer': "ONE",
        'rideDate': "2021/12/09",
        'startOrEndTime': "true",
        'startTime': "",
        'endTime': "",
        'trainTypeList': "ALL",
        '_isQryEarlyBirdTrn': "on",
        'query': "查詢"
        }
data['rideDate'] = day
data['startTime'] = time
data['endTime'] = end
data['startStation'] = getstations_input(st)
data['endStation'] = getstations_input(ed)
print(data)

'''wab crawler'''

url = 'https://tip.railway.gov.tw//tra-tip-web/tip/tip001/tip112/querybytime'
session = requests.Session()

s = session.get(url)
bs = BS(s.text, 'html.parser')

token = bs.find('form', {'id':"queryForm"}).input
#print(token['value'])

data['_csrf']=token['value']

r = session.post(url, data=data)
try:
    
    bs = BS(r.text, 'html.parser')

    trains= bs.find('div', {'class':"search-trip"}).find_all('tr', {'class':"trip-column"})

    spendtime = []
    for train in trains:
        result = train.find_all('td', text=re.compile('\d{2}.*'))
        departure_time = result[0].get_text()
        arrival_time =  result[1].get_text()
        tarvel_time = result[2].get_text()
        print('出發時間: ' + departure_time,
              '抵達時間: ' + arrival_time,
              '行駛時間: ' + tarvel_time,
              sep=' ', end='\n')
except:
    
    print('We only can search the train schedule in one day.')
    print('Maybe you should research after 00:00.')
print(t.time()-start)
    #can get arrive, departure and interval time.

# #python Day_28_json_project.py Taipei, TW  